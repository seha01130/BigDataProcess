#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
s = wb.active

list = []

row = 2
while True:
    if s.cell(row=row,column=1).value == None:
        break
    total = s.cell(row=row,column=3).value*0.3 + s.cell(row=row,column=4).value*0.35 + s.cell(row=row,column=5).value*0.34 + s.cell(row=row,column=6).value*1
    #print(total)
    if isinstance(total, float):
        total = "%.2f" % total
        total = float(total)
    else:
        total = int(total)
    s.cell(row=row,column=7, value = total)
    list.append(total)
    row += 1

# list.sort()
# list.reverse()
# print(type(list[0]))
# print(list)

list = sorted(list, reverse=True)
studentNum = len(list)

A_Cut = int(studentNum * 0.3)
A_Cut_Score = list[A_Cut-1] #A_Cut_Score = A0 문닫고 받는 사람의 점수
# print(A_Cut_Score)

APlus_Cut = A_Cut // 2
APlus_Cut_Score = list[APlus_Cut-1] #APlus_Cut_Score = A+ 문닫고 받는 사람의 점수
# print(APlus_Cut_Score)

B_Cut = int(studentNum * 0.7)
B_Cut_Score = list[B_Cut-1] #B_Cut_Score = B0 문닫고 받는 사람의 점수
# print(B_Cut_Score)

BPlus_Cut = B_Cut // 2
BPlus_Cut_Score = list[BPlus_Cut-1] #BPlus_Cut_Score = B+ 문닫고 받는 사람의 점수
# print(BPlus_Cut_Score)

count = 0
for i in range(B_Cut, studentNum):
    if list[i] >= 40:
        count += 1
C_Cut_Score = list[B_Cut-1+count] #C_Cut_Score = C0 문닫고 받는 사람의 점수
CPlus_Cut_Score = list[B_Cut-1+(count//2)] #CPlus_Cut_Score = C+ 문닫고 받는 사람의 점수

for i in range(2, len(list)+2):
    if s.cell(row=i,column=7).value < 40:
        s.cell(row=i,column=8, value = 'F')
    elif s.cell(row=i,column=7).value >= APlus_Cut_Score: #A+
        s.cell(row=i,column=8, value = 'A+')
    elif s.cell(row=i,column=7).value >= A_Cut_Score: #A
        s.cell(row=i,column=8, value = 'A0')
    elif s.cell(row=i,column=7).value >= BPlus_Cut_Score: #B+
        s.cell(row=i,column=8, value = 'B+')
    elif s.cell(row=i,column=7).value >= B_Cut_Score: #B0
        s.cell(row=i,column=8, value = 'B0')
    elif s.cell(row=i,column=7).value >= CPlus_Cut_Score: #C+
        s.cell(row=i,column=8, value = 'C+')
    elif s.cell(row=i,column=7).value >= C_Cut_Score: #C0
        s.cell(row=i,column=8, value = 'C0')

wb.save('student.xlsx')
wb.close()